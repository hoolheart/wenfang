"""
Tools to handle excel files conveniencely
"""
import pandas as pd
import openpyxl
from typing import (
    Dict,
    Iterable,
    Tuple,
    Union,
    Optional,
    Callable,
)
from openpyxl.utils.dataframe import dataframe_to_rows

"""
fetch all sheets from one excel file into dictionary of dataframes
"""
def get_sheets_to_dataframes(
        filename: str,
        filter: Optional[Callable[[int, str], bool]] = None,
) -> Dict[str, pd.DataFrame]:
    wb: openpyxl.Workbook = openpyxl.load_workbook(filename, read_only=True)
    sheets = {}
    for idx, ws_name in enumerate(wb.sheetnames):
        if isinstance(filter, Callable) and not filter(idx, ws_name):
            continue
        sheets[ws_name] = pd.read_excel(filename, ws_name)
    return sheets

def save_dataframes_to_excel(
        filename: str,
        dataframes: Union[Iterable[Tuple[str, pd.DataFrame]], 
                          Dict[str, pd.DataFrame]],
        save_index: bool = False,
) -> int:
    wb = openpyxl.Workbook()
    for ws_name in wb.sheetnames:
        wb.remove(wb[ws_name]) # remove all existing sheet
    seq = dataframes
    if isinstance(dataframes, Dict):
        seq = dataframes.items()
    count = 0
    if isinstance(seq, Iterable):
        for pair in seq:
            if isinstance(pair, Tuple) and len(pair) == 2 and \
                    isinstance(pair[0], str) and \
                    isinstance(pair[1], pd.DataFrame):
                sheet = wb.create_sheet(pair[0], count)
                for row in dataframe_to_rows(
                        pair[1], index=save_index, header=True):
                    sheet.append(row)
                count = count + 1
    wb.save(filename=filename)
    return count
